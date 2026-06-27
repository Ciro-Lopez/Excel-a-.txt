from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("Prueba.xlsx")
ws = wb.active
j=0
long=0

valColumns=[]

cantColumns=int(ws.max_column)

for i in range(cantColumns):
    valColumns.append(input("¿Qué columna quieres que sea la fila número " + str(i) + "?\n"))

i=0

for fila in ws.iter_rows(values_only=True):
    for column in fila:
        with open('archivo.txt', 'a', encoding='utf-8') as archivo:
            if j==0:
                for i in range(int(valColumns[j])):
                    archivo.write(" ")
                archivo.write(str(column))
                j+=1
            long+=len(str(column))
            if j!=1:
                archivo.write(str(column))
            
            if j<cantColumns:
                cantEsp=int(valColumns[j])-int(long)
                cantEsp-=1
                long+=cantEsp
                print(cantEsp)
                if cantEsp<0:
                    print("Hay una fila que no entra en la cantidad de espacios dados")
                    exit()
                for i in range(cantEsp):
                    archivo.write(" ")
        j+=1
    j=0
    long=0
    with open('archivo.txt', 'a', encoding='utf-8') as archivo:
        archivo.write("\n")